import openpyxl
from openpyxl import load_workbook
excel=load_workbook(r'C:\Users\DELL\Desktop\debian_data.xlsx')
sheet=excel.active
ww1=excel[excel.sheetnames[0]]
#截取想要的数据放到datalist列表中
for row in sheet:
    sheet.append(row)



yearlist=[]
for i in range(1,4981):
    i = str(i)
    if ww1['E'+i].value == "Yes":
        yearlist.append(ww1['A'+i].value[-5:-1])
        print(ww1['A'+i].value[-5:-1], end=" ")
    else:
        continue
print(len(yearlist))
#a= {}
b= {}
for i in yearlist:
    if yearlist.count(i) > 1:
     #   a[i] = datalist.count(i)
        b[i] = yearlist.count(i)

b1 = sorted(b.items(),key=lambda x:x[0],reverse=False)
print(b1)
datalist=[]
#for j in range(len(b1)):
#2-18 19-61 62-
#2-d[0]+1 d[0]+min_row-max_row+d[1]
    


for row in sheet.iter_rows(min_row=2, min_col=5, max_row=18, max_col=5):
    for cell in row:
        print(cell.value, end=" ")
        datalist.append(cell.value)    
print(len(datalist))
from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)

book = Workbook()
sheet = book.active

for row in b1:
    sheet.append(row)

data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=len(b1))
categs = Reference(sheet, min_col=1, min_row=1, max_row=len(b1))

chart = BarChart()
chart.add_data(data=data)
chart.set_categories(categs)

chart.legend = None
chart.y_axis.majorGridlines = None  #关闭图例和主要网格线
chart.varyColors = True  #不同颜色
chart.title = "Vulnerable"  #标题

sheet.add_chart(chart, "A30")    

book.save("vulnerable.xlsx")




