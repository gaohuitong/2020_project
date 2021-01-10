import openpyxl
from openpyxl import load_workbook
excel=load_workbook(r'C:\Users\DELL\Desktop\debian_data.xlsx')
sheet=excel.active
#截取想要的数据放到datalist列表中
for row in sheet:
    sheet.append(row)
datalist=[]
for row in sheet.iter_rows(min_row=2, min_col=1, max_row=4981, max_col=1):
    for cell in row:
        print(cell.value[-9:-5], end=" ")
        datalist.append(cell.value[-9:-5])
print(len(datalist))
#统计重复字符串，对返回的结果按照字典的值从大到小排序
a= {}
for i in datalist:
    if datalist.count(i) > 1:
        a[i] = datalist.count(i)

a1 = sorted(a.items(),key=lambda x:x[1],reverse=True)
print(a1)
#将a1列表中的元组存入新的Excel中进行作图
from openpyxl import Workbook
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)

book = Workbook()
sheet = book.active

for row in a1:
    sheet.append(row)

data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=len(a1))
categs = Reference(sheet, min_col=1, min_row=1, max_row=len(a1))

chart = BarChart()
chart.add_data(data=data)
chart.set_categories(categs)

chart.legend = None
chart.y_axis.majorGridlines = None  #关闭图例和主要网格线
chart.varyColors = True  #不同颜色
chart.title = "Month"  #标题

sheet.add_chart(chart, "A20")    

book.save("month.xlsx")
